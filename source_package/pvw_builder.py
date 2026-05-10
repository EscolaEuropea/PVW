"""
PVW.one v2.0 — Shared archetype builder library
Generates ARCHETYPE_*.xlsx + archetype_*.schema.json from a Python data spec.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
from datetime import date

# Palette
NAVY = "0F1B3D"
TEAL = "0FA3B1"
HEADER_BG = "0F1B3D"
SUBHEADER_BG = "1A2A5C"
DEFAULT_S_BG = "E8F4F6"
DEFAULT_M_BG = "D1E9ED"
DEFAULT_L_BG = "B8DCE2"
META_BG = "F4F1EA"

PVW_VERSION = "1.0.0"

# ─────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────
def title_banner(ws, row, title, subtitle, span=8):
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(2, span+1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 28
    if subtitle:
        c2 = ws.cell(row=row+1, column=1, value=subtitle)
        c2.font = Font(name="Calibri", size=10, color="C5CCE0", italic=True)
        c2.fill = PatternFill("solid", fgColor=NAVY)
        for col in range(2, span+1):
            ws.cell(row=row+1, column=col).fill = PatternFill("solid", fgColor=NAVY)
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=span)
        ws.row_dimensions[row+1].height = 18

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def data_row(ws, row, values, default_cols=None):
    default_cols = default_cols or {}
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name="Calibri", size=10, color="000000")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if i in default_cols:
            tint = default_cols[i]
            bg = {"S": DEFAULT_S_BG, "M": DEFAULT_M_BG, "L": DEFAULT_L_BG}[tint]
            c.fill = PatternFill("solid", fgColor=bg)

def configure_print(ws):
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_setup.orientation = "landscape"

# ─────────────────────────────────────────────────────────────────
# Main builder
# ─────────────────────────────────────────────────────────────────
def build_archetype(spec, out_dir):
    """spec: dict with archetype_id, name, category, definition,
            tier_S/M/L description, identity, equipment, relations, sources, defaults_M.
    """
    aid = spec["archetype_id"]
    name = spec["name"]
    category = spec["category"]
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ─────────────────────────────────────────────
    # 1_README
    # ─────────────────────────────────────────────
    ws = wb.create_sheet("1_README")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80
    title_banner(ws, 1, f"PVW.one ARCHETYPE — {name}",
                 f"archetype_id: {aid} · version: {PVW_VERSION} · category: {category}")
    
    rows = [
        ("ARCHETYPE METADATA", ""),
        ("archetype_id", aid),
        ("archetype_name", name),
        ("archetype_category", category),
        ("archetype_version", PVW_VERSION),
        ("created", date.today().isoformat()),
        ("schema_file", f"archetype_{aid}.schema.json"),
        ("", ""),
        ("DEFINITION", ""),
        ("", spec["definition"]),
        ("", ""),
        ("SIZE TIERS (the 3 default profiles)", ""),
        ("S — Small", spec["tier_S"]),
        ("M — Medium", spec["tier_M"]),
        ("L — Large", spec["tier_L"]),
        ("", ""),
        ("HOW TO USE", ""),
        ("",
         "(1) This is the ARCHETYPE TEMPLATE: it defines field_ids, units, types, and S/M/L defaults. "
         "(2) To create an INSTANCE, copy the structure into a new INSTANCE_*.xlsx, declare its size tier, "
         "inherit defaults, overwrite cells with real data as available. "
         "(3) Each instance cell carries metadata: fidelity (L0/L1/L2), source, last_updated. "
         "(4) The companion JSON Schema defines the formal contract for BBDD migration."),
        ("", ""),
        ("FIDELITY LEVELS (used by instances)", ""),
        ("L0 — Default", "Inherited from this archetype's S/M/L column. No real data."),
        ("L1 — Public refined", "Verified data from public sources."),
        ("L2 — Operator twin", "Data provided directly by the operator under collaboration."),
    ]
    for i, (k, v) in enumerate(rows, 4):
        c1 = ws.cell(row=i, column=1, value=k)
        c1.font = Font(name="Calibri", size=11, bold=(k.isupper() and k != ""))
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c2 = ws.cell(row=i, column=2, value=v)
        c2.font = Font(name="Calibri", size=10)
        c2.alignment = Alignment(vertical="top", wrap_text=True)
    
    # ─────────────────────────────────────────────
    # 2_Identity_Schema
    # ─────────────────────────────────────────────
    ws = wb.create_sheet("2_Identity_Schema")
    title_banner(ws, 1, "Identity — Schema & defaults",
                 "Each row defines one field. S/M/L columns provide tier-typical defaults.")
    headers = ["field_id", "label", "unit", "type", "default_S", "default_M", "default_L", "notes"]
    widths = [32, 30, 10, 10, 16, 16, 18, 32]
    header_row(ws, 4, headers, widths)
    for i, row in enumerate(spec["identity"]):
        data_row(ws, 5+i, list(row), default_cols={5: "S", 6: "M", 7: "L"})
    configure_print(ws)
    
    # ─────────────────────────────────────────────
    # 3_Equipment_Schema
    # ─────────────────────────────────────────────
    ws = wb.create_sheet("3_Equipment_Schema")
    title_banner(ws, 1, "Equipment — Schema & defaults", "Physical assets and operational capacity.")
    header_row(ws, 4, headers, widths)
    for i, row in enumerate(spec["equipment"]):
        data_row(ws, 5+i, list(row), default_cols={5: "S", 6: "M", 7: "L"})
    configure_print(ws)
    
    # ─────────────────────────────────────────────
    # 4_Relations_Schema
    # ─────────────────────────────────────────────
    ws = wb.create_sheet("4_Relations_Schema")
    title_banner(ws, 1, "Relations — Counterparty types & typical volumes",
                 "Each row = one type of counterparty. Specific actors filled per instance.")
    headers_r = ["counterparty_type_id", "label", "direction", "frequency",
                 "vol_default_S", "vol_default_M", "vol_default_L", "notes"]
    widths_r = [32, 30, 14, 14, 14, 14, 16, 30]
    header_row(ws, 4, headers_r, widths_r)
    for i, row in enumerate(spec["relations"]):
        data_row(ws, 5+i, list(row), default_cols={5: "S", 6: "M", 7: "L"})
    configure_print(ws)
    
    # ─────────────────────────────────────────────
    # 5_Carbon_Schema
    # ─────────────────────────────────────────────
    ws = wb.create_sheet("5_Carbon_Schema")
    title_banner(ws, 1, "Carbon — Emission sources, scenarios, defaults",
                 "Schema: each row = one (source × scenario × year) cell.")
    headers_c = ["emission_source_id", "label", "scope", "scenario", "year",
                 "default_S_tCO2e", "default_M_tCO2e", "default_L_tCO2e"]
    widths_c = [28, 32, 8, 16, 8, 14, 14, 14]
    header_row(ws, 4, headers_c, widths_c)
    
    sources = spec["sources"]
    scenarios = ["BAU", "FuelEU+AFIR", "Accelerated"]
    defaults_M = spec["defaults_M"]
    
    row = 5
    for src_id, label, scope in sources:
        for scenario in scenarios:
            for yi, year in enumerate(range(2025, 2031)):
                m_def = defaults_M[src_id][scenario][yi]
                s_def = round(m_def * 0.30)
                l_def = round(m_def * 2.50)
                data_row(ws, row, [src_id, label, scope, scenario, year, s_def, m_def, l_def],
                         default_cols={6: "S", 7: "M", 8: "L"})
                row += 1
    configure_print(ws)
    
    # ─────────────────────────────────────────────
    # 99_Field_Catalog
    # ─────────────────────────────────────────────
    ws = wb.create_sheet("99_Field_Catalog")
    title_banner(ws, 1, "Field Catalog — All field_ids in this archetype",
                 "Reference for instances and BBDD migration.")
    header_row(ws, 4, ["field_id", "sheet", "label", "unit", "type"],
               [36, 24, 32, 10, 12])
    cat = []
    for f in spec["identity"]:
        cat.append((f[0], "2_Identity_Schema", f[1], f[2], f[3]))
    for f in spec["equipment"]:
        cat.append((f[0], "3_Equipment_Schema", f[1], f[2], f[3]))
    for f in spec["relations"]:
        cat.append((f[0], "4_Relations_Schema", f[1], "—", "counterparty"))
    for src_id, label, scope in spec["sources"]:
        cat.append((src_id, "5_Carbon_Schema", label, "t CO2e", "emission_source"))
    for i, row in enumerate(cat):
        data_row(ws, 5+i, list(row))
    configure_print(ws)
    
    # Save
    out_xlsx = f"{out_dir}/ARCHETYPE_{aid}.xlsx"
    wb.save(out_xlsx)
    
    # ─────────────────────────────────────────────
    # JSON Schema
    # ─────────────────────────────────────────────
    schema = build_json_schema(spec)
    out_json = f"{out_dir}/archetype_{aid}.schema.json"
    with open(out_json, "w") as f:
        json.dump(schema, f, indent=2, ensure_ascii=False)
    
    return out_xlsx, out_json, len(cat)


def f_to_jsontype(t):
    return {"string": "string", "integer": "integer", "number": "number",
            "boolean": "boolean", "enum": "string"}.get(t, "string")


def build_json_schema(spec):
    aid = spec["archetype_id"]
    name = spec["name"]
    
    VOBJ = {
        "type": "object",
        "required": ["value", "fidelity"],
        "properties": {
            "value": {},
            "fidelity": {"type": "string", "enum": ["L0", "L1", "L2"]},
            "source": {"type": "string"},
            "last_updated": {"type": "string"}
        },
        "additionalProperties": False
    }
    
    def field_props(field_list):
        out = {}
        for f in field_list:
            fid, label, unit, ftype = f[0], f[1], f[2], f[3]
            d = dict(VOBJ)
            d["title"] = label
            if unit and unit != "—":
                d["x_unit"] = unit
            d["x_type"] = ftype
            out[fid] = d
        return out
    
    schema = {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "$id": f"https://pvw.one/schemas/archetype/{aid}/{PVW_VERSION}",
        "title": f"{name} — Archetype Schema",
        "description": f"Defines the structure of any {name} instance in PVW.one. Version {PVW_VERSION}.",
        "version": PVW_VERSION,
        "archetype_id": aid,
        "archetype_category": spec["category"],
        "type": "object",
        "required": ["instance_metadata", "identity", "equipment", "relations", "carbon"],
        "properties": {
            "instance_metadata": {
                "type": "object",
                "required": ["instance_id", "archetype_ref", "port_unlocode", "size_tier", "last_updated"],
                "properties": {
                    "instance_id": {"type": "string"},
                    "instance_name": {"type": "string"},
                    "archetype_ref": {"type": "string"},
                    "port_unlocode": {"type": "string"},
                    "port_name": {"type": "string"},
                    "country_iso2": {"type": "string", "minLength": 2, "maxLength": 2},
                    "size_tier": {"type": "string", "enum": ["S", "M", "L"]},
                    "active_carbon_scenario": {"type": "string", "enum": ["BAU", "FuelEU+AFIR", "Accelerated"]},
                    "created": {"type": "string"},
                    "last_updated": {"type": "string"},
                    "schema_version": {"type": "string"}
                }
            },
            "identity": {
                "type": "object",
                "additionalProperties": False,
                "properties": field_props(spec["identity"])
            },
            "equipment": {
                "type": "object",
                "additionalProperties": False,
                "properties": field_props(spec["equipment"])
            },
            "relations": {
                "type": "object",
                "description": "For each counterparty_type_id defined in the archetype, an object with actors and metadata.",
                "additionalProperties": True
            },
            "carbon": {
                "type": "object",
                "required": ["active_scenario", "emissions"],
                "properties": {
                    "active_scenario": {"type": "string", "enum": ["BAU", "FuelEU+AFIR", "Accelerated"]},
                    "emissions": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "required": ["emission_source_id", "scope", "scenario", "year", "value_tCO2e", "fidelity"],
                            "properties": {
                                "emission_source_id": {"type": "string"},
                                "scope": {"type": "string", "enum": ["scope_1", "scope_2", "scope_1+2", "scope_3"]},
                                "scenario": {"type": "string", "enum": ["BAU", "FuelEU+AFIR", "Accelerated"]},
                                "year": {"type": "integer", "minimum": 2020, "maximum": 2050},
                                "value_tCO2e": {"type": "number", "minimum": 0},
                                "fidelity": {"type": "string", "enum": ["L0", "L1", "L2"]},
                                "source": {"type": "string"}
                            }
                        }
                    }
                }
            }
        },
        "x_pvw_fidelity_levels": {
            "L0": "Default — inherited from archetype S/M/L. No real data.",
            "L1": "Public refined — verified from public sources.",
            "L2": "Operator twin — provided by the operator under collaboration."
        }
    }
    return schema


def make_carbon_defaults(sources_with_baselines, scenario_modifiers=None):
    """sources_with_baselines: list of (source_id, baseline_2025_tCO2e_M)
    Returns defaults_M dict source -> {scenario: [2025...2030]}
    
    Default decay curves:
      BAU:         -1% per year
      FuelEU+AFIR: -6, -8, -10, -12, -15 vs prior year
      Accelerated: -15, -20, -25, -30, -35 vs prior year
    """
    BAU_factors = [1.00, 0.99, 0.98, 0.97, 0.96, 0.95]
    FEU_factors = [1.00, 0.94, 0.86, 0.77, 0.68, 0.58]
    ACC_factors = [1.00, 0.85, 0.68, 0.51, 0.36, 0.23]
    
    out = {}
    for src_id, baseline in sources_with_baselines:
        out[src_id] = {
            "BAU":         [round(baseline * f) for f in BAU_factors],
            "FuelEU+AFIR": [round(baseline * f) for f in FEU_factors],
            "Accelerated": [round(baseline * f) for f in ACC_factors],
        }
    return out
